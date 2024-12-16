import os
from pathlib import Path

import pandas as pd
import xlwings as xw
from faker import Faker
from openpyxl import load_workbook
from pofile import get_files
from pofile import mkdir
from poprogress import simple_progress
from tqdm import tqdm

# 忽略waring警告
from poexcel.lib import pandas_mem
from poexcel.lib.excel import SplitExcel


class MainExcel():
    def __init__(self):
        self.app = "Excel.Application"

    def fake2excel(self, columns, rows, path, language):
        """
        @Author & Date  : CoderWanFeng 2022/5/13 0:12
        @Desc  : columns:list，每列的数据名称，默认是名称
                rows：多少行，默认是1
                language：什么语言，可以填english，默认是中文
                path：输出excel的位置，有默认值
        @Ref  : 可以fake的数据类型有：https://mp.weixin.qq.com/s/xVwEjXu58WovgSi4ZTtVQw
        """
        if rows == 0:
            pd.DataFrame().to_excel(str(Path(path).absolute()))
        else:
            # 可以选择英语
            language = 'en_US' if language.lower() == 'english' else 'zh_CN'
            fake = Faker(language)
            excel_dict = {}
            for column in simple_progress(columns, desc=f'columns'):
                excel_dict[column] = list()
                for _ in simple_progress(range(0, rows), desc='rows'):
                    excel_dict[column].append(eval('fake.{func}()'.format(func=column)))
            # 用pandas，将模拟数据，写进excel里面
            res_excel_file = pd.ExcelWriter(str(Path(path).absolute()))
            res_data = pd.DataFrame(excel_dict)
            res_data = pandas_mem.reduce_pandas_mem_usage(res_data)
            res_data.to_excel(res_excel_file, index=False)
            # writer.save()
            res_excel_file.close()

    def merge2excel(self, dir_path, output_file, xlsxSuffix=".xlsx"):
        """
        :param dir_path: 存放excel文件的位置
        :param output_file: 输出合并后excel文件的位置
        :return: 没有返回值
        """
        abs_output_path = Path(output_file).absolute()
        mkdir(abs_output_path.parent)  # 创建输出路径
        if not output_file.endswith(xlsxSuffix):
            raise Exception(f'您自定义的输出文件名，不是以{xlsxSuffix}结尾的')
        file_path_dict = self.getfile(dir_path)  # excel文件所在的文件夹
        try:
            writer = pd.ExcelWriter(output_file)  # 合并后的excel名称
        except PermissionError:
            raise Exception(f'小可爱，你的输出文件，是不是上次打开了没关闭呀？这是你自己指定的输出文件名称：{output_file}')
        for file, path in file_path_dict.items():
            if file.endswith("xlsx"):
                df = pd.read_excel(path)
            if file.endswith("csv"):
                df = pd.read_csv(path)
            df.to_excel(writer, sheet_name=file.split('.')[0], index=False)
        print(f'您指定的Excel文件已经合并完毕，合并后的文件名是{output_file}')
        writer._save()

    def getfile(self, dirpath):
        path = Path(dirpath)
        file_path_dict = {}
        for root, dirs, files in os.walk(dirpath):
            for file in files:
                if file.endswith("xlsx") or file.endswith("csv"):
                    file_path_dict[file] = (path / file)
        return file_path_dict

    def sheet2excel(self, file_path, output_path: str):
        # 先读取一次文件，获取sheet表的名称

        origin_excel = load_workbook(filename=file_path)  # 读取原excel文件
        origin_sheet_names = origin_excel.sheetnames  # 获取sheet的名称
        print(f'一共有{len(origin_sheet_names)}个sheet，名称分别为：{origin_sheet_names}')
        print('拆分开始')

        if len(origin_sheet_names) > 1:  # 如果sheetnames小于1，报错：该文件不需要拆分

            for j in tqdm(range(len(origin_sheet_names))):

                wb = load_workbook(filename=file_path)  # 再读取一次文件，由于每次删除后需要保存一次，所以不能与上一次一样
                sheet = wb[origin_sheet_names[j]]
                wb.copy_worksheet(sheet)

                new_filename = Path(output_path).joinpath(origin_sheet_names[j] + '.xlsx')  # 新建一个sheet命名的excel文件

                for i in tqdm(range(len(origin_sheet_names))):
                    sheet1 = wb[origin_sheet_names[i]]
                    wb.remove(sheet1)

                wb.save(filename=new_filename)

                # 由于使用copy_worksheet后，sheet表名有copy字段，这里做个调整

                new = load_workbook(filename=new_filename)
                news = new.active
                news.title = origin_sheet_names[j]
                new.save(filename=new_filename)
            print('拆分结束')
        else:
            raise Exception(f"你的文件只有一个sheet，难道还要拆分吗？我做不到啊~~~，你的文件名{file_path}")

    def merge2sheet(self, dir_path, output_excel_name, output_sheet_name: str = None):
        """
        将指定目录下的多个文件合并到一个Excel表格中。

        参数：
        dir_path (str): 指定目录的路径。
        output_excel_name (str): 输出的Excel文件名。
        output_sheet_name (str, optional): 输出的Excel表格名，默认为None。

        返回：
        None
        """

        def merge_excel_by_sheet(path, file_list, sheet_name):
            """
            :param file_path: excel文件的路径
            :param output_path: 输出excel文件的路径
            :return:
            """
            # 读取excel文件
            df_list = []
            for excel_path in file_list:  # 依次取出Excel文件
                if excel_path.endswith("xlsx") or excel_path.endswith("xls"):
                    excel_path = Path(path) / excel_path
                    df_list.append(pd.read_excel(excel_path, sheet_name=sheet_name))
            res = pd.concat(df_list)
            return res

        for root, dirs, files in os.walk(dir_path):
            path = Path(dir_path)
            output_excel_path = Path(output_excel_name).parent.absolute()
            mkdir(output_excel_path)
            writer = pd.ExcelWriter(output_excel_name + '.xlsx', mode='w', engine='openpyxl')
            df_list = []
            sheet_merge_dic = {}
            for file in files:  # 依次取出Excel文件
                if file.endswith("xlsx") or file.endswith("xls"):
                    excel_path = (path / file)
                    if output_sheet_name != None:
                        sheet_merge_dic[output_sheet_name] = merge_excel_by_sheet(path, files, output_sheet_name)
                    else:
                        for sheet_name in pd.ExcelFile(excel_path).sheet_names:
                            sheet_merge_dic[sheet_name] = merge_excel_by_sheet(path, files, sheet_name)
                    break
            for sheet_name, df in sheet_merge_dic.items():
                print(sheet_name)
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            writer._save()
            writer.close()

    def find_excel_data(self, search_key, target_dir):
        """
        检索指定目录下的excel文件和过滤
        参数：
            search_key：检索的关键词
            target_dir：目标文件夹
        """
        print('该方法已过期，请调用：search4file.find_excel_data(search_key, target_dir)')

    def split_excel_by_column(self, filepath, column, worksheet_name):
        SplitExcel.split_excel_by_column(filepath, column, worksheet_name)

    def excel2pdf(self, excel_path, pdf_path, sheet_id):
        # 获取绝对路径
        abs_excel_path = str(Path(excel_path).absolute())
        # 获取指定后缀的Excel文件列表（.xls）
        input_excel_path_list1 = get_files(abs_excel_path, suffix='.xls')
        # 获取指定后缀的Excel文件列表（.xlsx）
        input_excel_path_list2 = get_files(abs_excel_path, suffix='.xlsx')
        # 将两个列表合并
        input_excel_path_list1.extend(input_excel_path_list2)
        # 获取PDF文件的绝对路径
        output_pdf_path = Path(pdf_path).absolute()
        # 创建PDF文件的输出目录
        mkdir(output_pdf_path)
        # 遍历Excel文件列表
        for excel_file in input_excel_path_list1:
            with xw.App() as app:  # 下列来源：https://www.qiniu.com/qfans/qnso-57724345#comments
                # 设置Excel应用程序不可见
                app.visible = False
                # 初始化新的Excel工作簿
                # Initialize new excel workbook
                book = app.books.open(str(excel_file))
                # 获取指定sheet_id的工作表
                sheet = book.sheets[sheet_id]
                # 构造PDF文件的路径
                # Construct path for pdf file
                pdf_path_name = os.path.join(str(output_pdf_path), Path(excel_file).stem + '.pdf')
                # 将工作表保存为PDF文件
                sheet.to_pdf(path=pdf_path_name, show=False)

    def count4page(self, input_path):
        """
        统计Excel文件打印的页数
        :author Cai-cy
        :param input_path:
        :return:
        """
        # # 指定文件夹路径
        # # 打开 Excel 应用程序
        # excel = win32com.client.Dispatch(self.app)
        #
        # # 遍历文件夹下的所有文件
        # for file_name in os.listdir(input_path):
        #     # 判断文件是否是 Excel 文件
        #     if file_name.endswith(".xlsx") or file_name.endswith(".xls"):
        #         # 打开 Excel 文件
        #         file_path = os.path.join(input_path, file_name)
        #         workbook = excel.Workbooks.Open(file_path)
        #         # 获取 Excel 文件的打印页数
        #         page_count = workbook.ActiveSheet.PageSetup.Pages.Count
        #         # 输出 Excel 文件的打印页数
        #         print(f"{file_name}: {page_count}页")
        #         # 关闭 Excel 文件
        #         workbook.Close()
        #
        # # 关闭 Excel 应用程序
        # excel.Quit()
        pass
